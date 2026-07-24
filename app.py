"""
THE CHOICE HUB – FULL E‑COMMERCE (Google OAuth + Email Password Reset)
All features: Customer, Admin, Seller, Cart, Checkout, Coupons, Offers,
Reviews, Wishlist, Pincode Delivery, Order Tracking, Excel import, etc.
"""
import os
import uuid
import random
from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
import requests

load_dotenv()  # Local development ke liye

# ---------- APP SETUP ----------
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///choicehub.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Upload folder
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------- MAIL SETUP ----------
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() in ['true', '1']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER')

mail = Mail(app)

# ---------- GOOGLE OAUTH ----------
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.environ.get('GOOGLE_CLIENT_ID'),
    client_secret=os.environ.get('GOOGLE_CLIENT_SECRET'),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'}
)

# ---------- DB & LOGIN ----------
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ---------- MODELS ----------
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(15), unique=True, nullable=True)
    email = db.Column(db.String(100), unique=True, nullable=True)
    password_hash = db.Column(db.String(200), nullable=True)
    google_id = db.Column(db.String(100), unique=True, nullable=True)
    profile_pic = db.Column(db.String(200), nullable=True)
    role = db.Column(db.String(20), default='customer')
    referral_code = db.Column(db.String(20), unique=True, nullable=True)
    wishlist = db.Column(db.Text, nullable=True)
    reset_token = db.Column(db.String(100), unique=True, nullable=True)        # ← नया
    reset_token_expiry = db.Column(db.DateTime, nullable=True)                 # ← नया
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    orders = db.relationship('Order', backref='customer', lazy=True)
    reviews = db.relationship('Review', backref='user', lazy=True)

class Seller(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    store_name = db.Column(db.String(100), nullable=True)
    commission_rate = db.Column(db.Float, default=10.0)
    total_earned = db.Column(db.Float, default=0.0)
    user = db.relationship('User', backref='seller_profile')
    products = db.relationship('Product', backref='seller', lazy=True)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    slug = db.Column(db.String(100), unique=True, nullable=False)
    image = db.Column(db.String(200), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)
    children = db.relationship('Category', backref=db.backref('parent', remote_side=[id]), lazy=True)
    products = db.relationship('Product', backref='category_ref', lazy=True)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    seller_id = db.Column(db.Integer, db.ForeignKey('seller.id'), nullable=True)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(200), unique=True, nullable=False)
    short_description = db.Column(db.String(300), nullable=True)
    description = db.Column(db.Text, nullable=True)
    features = db.Column(db.Text, nullable=True)
    specifications = db.Column(db.Text, nullable=True)
    brand = db.Column(db.String(100), nullable=True)
    sku = db.Column(db.String(100), unique=True, nullable=False)
    cost_price = db.Column(db.Float, nullable=True)
    selling_price = db.Column(db.Float, nullable=False)
    mrp = db.Column(db.Float, nullable=False)
    discount_percent = db.Column(db.Float, default=0.0)
    stock = db.Column(db.Integer, default=0)
    weight = db.Column(db.Float, nullable=True)
    dimensions = db.Column(db.String(100), nullable=True)
    material = db.Column(db.String(100), nullable=True)
    care_instructions = db.Column(db.Text, nullable=True)
    warranty = db.Column(db.String(100), nullable=True)
    return_policy = db.Column(db.Text, nullable=True)
    delivery_time = db.Column(db.String(50), nullable=True)
    images = db.Column(db.Text, nullable=True)
    video_url = db.Column(db.String(300), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_featured = db.Column(db.Boolean, default=False)
    is_bestseller = db.Column(db.Boolean, default=False)
    is_trending = db.Column(db.Boolean, default=False)
    is_new = db.Column(db.Boolean, default=False)
    free_delivery = db.Column(db.Boolean, default=False)
    cod_available = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def discount(self):
        if self.mrp and self.mrp > 0 and self.selling_price < self.mrp:
            return int(((self.mrp - self.selling_price) / self.mrp) * 100)
        return 0

    def savings(self):
        return self.mrp - self.selling_price if self.mrp > self.selling_price else 0

    def image_list(self):
        if self.images:
            return [img.strip() for img in self.images.split(',') if img.strip()]
        return []

    def main_image(self):
        imgs = self.image_list()
        return imgs[0] if imgs else None

    def avg_rating(self):
        reviews = Review.query.filter_by(product_id=self.id, is_approved=True).all()
        if not reviews:
            return 0
        return round(sum(r.rating for r in reviews) / len(reviews), 1)

    def review_count(self):
        return Review.query.filter_by(product_id=self.id, is_approved=True).count()

class Cart(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    is_gift = db.Column(db.Boolean, default=False)
    user = db.relationship('User', backref='cart_items')
    product = db.relationship('Product')

class Coupon(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    discount_type = db.Column(db.String(20), default='percentage')
    discount_value = db.Column(db.Float, nullable=False)
    min_order_amount = db.Column(db.Float, default=0)
    max_discount_amount = db.Column(db.Float, nullable=True)
    expiry_date = db.Column(db.DateTime, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_number = db.Column(db.String(50), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    coupon_id = db.Column(db.Integer, db.ForeignKey('coupon.id'), nullable=True)
    referral_used = db.Column(db.String(20), nullable=True)
    total_amount = db.Column(db.Float, nullable=False)
    discount_amount = db.Column(db.Float, default=0.0)
    shipping_charge = db.Column(db.Float, default=0.0)
    net_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(30), default='Placed')
    payment_method = db.Column(db.String(50), default='COD')
    payment_status = db.Column(db.String(30), default='Pending')
    shipping_address = db.Column(db.Text, nullable=False)
    pincode = db.Column(db.String(10), nullable=True)
    delivery_date = db.Column(db.DateTime, nullable=True)
    tracking_number = db.Column(db.String(100), nullable=True)
    courier_name = db.Column(db.String(100), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    coupon = db.relationship('Coupon', backref='orders', lazy=True)
    items = db.relationship('OrderItem', backref='order', lazy=True)

class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    seller_id = db.Column(db.Integer, db.ForeignKey('seller.id'), nullable=True)
    quantity = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    product = db.relationship('Product')

class Review(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=True)
    rating = db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(100), nullable=True)
    comment = db.Column(db.Text, nullable=True)
    images = db.Column(db.Text, nullable=True)
    is_verified = db.Column(db.Boolean, default=False)
    is_approved = db.Column(db.Boolean, default=False)
    helpful_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Address(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(15), nullable=False)
    address_line1 = db.Column(db.String(200), nullable=False)
    address_line2 = db.Column(db.String(200), nullable=True)
    city = db.Column(db.String(100), nullable=False)
    state = db.Column(db.String(100), nullable=False)
    pincode = db.Column(db.String(10), nullable=False)
    country = db.Column(db.String(50), default='India')
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    locality = db.Column(db.String(100), nullable=True)
    landmark = db.Column(db.String(100), nullable=True)

class Banner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=True)
    image = db.Column(db.String(200), nullable=False)
    link = db.Column(db.String(300), nullable=True)
    position = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)

# ---------- DATABASE INIT (NO DEMO DATA) ----------
with app.app_context():
    db.create_all()
    # Ensure images column exists (legacy)
    try:
        db.session.execute("ALTER TABLE product ADD COLUMN images TEXT")
        db.session.commit()
    except:
        pass

    # Add reset_token and reset_token_expiry columns if not exist
    try:
        db.session.execute("ALTER TABLE user ADD COLUMN reset_token VARCHAR(100)")
        db.session.commit()
    except:
        pass
    try:
        db.session.execute("ALTER TABLE user ADD COLUMN reset_token_expiry DATETIME")
        db.session.commit()
    except:
        pass

    # ONLY create admin if NONE exists (no default categories/coupons/sellers)
    if not User.query.filter_by(role='admin').first():
        admin = User(
            name='Super Admin',
            phone='9999999999',
            email='admin@choicehub.com',
            password_hash=generate_password_hash('admin123'),
            role='admin',
            referral_code='ADMIN001'
        )
        db.session.add(admin)
        db.session.commit()
        print("✅ Admin user created (login: 9999999999 / admin123)")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ---------- HELPERS ----------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generate_order_number():
    return f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

def cart_total(user_id):
    items = Cart.query.filter_by(user_id=user_id).all()
    return sum(item.product.selling_price * item.quantity for item in items)

def estimate_delivery_date(pincode):
    if not pincode or len(pincode) != 6 or not pincode.isdigit():
        return None, None
    prefix = int(pincode[:2]) if pincode[:2].isdigit() else 0
    if prefix >= 11 and prefix <= 13:
        days = 2
    elif prefix >= 20 and prefix <= 28:
        days = 3
    elif prefix >= 30 and prefix <= 39:
        days = 4
    elif prefix >= 40 and prefix <= 49:
        days = 5
    elif prefix >= 50 and prefix <= 59:
        days = 5
    elif prefix >= 60 and prefix <= 69:
        days = 6
    elif prefix >= 70 and prefix <= 79:
        days = 4
    elif prefix >= 80 and prefix <= 89:
        days = 5
    else:
        days = 6
    delivery_date = datetime.utcnow() + timedelta(days=days)
    return delivery_date, days

# ---------- TOKEN GENERATION ----------
def generate_reset_token(email):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    return serializer.dumps(email, salt='password-reset-salt')

def confirm_reset_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(token, salt='password-reset-salt', max_age=expiration)
    except:
        return None
    return email

# ---------- GOOGLE AUTH ROUTES ----------
@app.route('/login/google')
def google_login():
    redirect_uri = url_for('google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/login/google/callback')
def google_callback():
    try:
        token = google.authorize_access_token()
        user_info = token.get('userinfo')
        if not user_info:
            user_info = google.get('https://www.googleapis.com/oauth2/v2/userinfo').json()
        
        email = user_info.get('email')
        google_id = user_info.get('sub')
        name = user_info.get('name', 'Google User')
        picture = user_info.get('picture')
        
        # Check if user exists
        user = User.query.filter_by(email=email).first()
        if not user:
            user = User.query.filter_by(google_id=google_id).first()
        
        if not user:
            # Create new user
            user = User(
                name=name,
                email=email,
                google_id=google_id,
                profile_pic=picture,
                role='customer',
                referral_code=f"GOOG_{uuid.uuid4().hex[:8].upper()}"
            )
            db.session.add(user)
            db.session.commit()
            flash('✅ Account created with Google!')
        
        login_user(user)
        flash(f'✅ Welcome {user.name}!')
        return redirect('/')
    except Exception as e:
        flash(f'❌ Google login failed: {str(e)}')
        return redirect('/login')

# ---------- REGULAR LOGIN/REGISTER ----------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        password = request.form.get('password')
        if User.query.filter_by(phone=phone).first():
            flash('Phone already registered!')
            return redirect('/register')
        if User.query.filter_by(email=email).first():
            flash('Email already registered!')
            return redirect('/register')
        hashed = generate_password_hash(password)
        user = User(
            name=name,
            phone=phone,
            email=email,
            password_hash=hashed,
            role='customer',
            referral_code=f"REF_{uuid.uuid4().hex[:8].upper()}"
        )
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please login.')
        return redirect('/login')
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        phone = request.form.get('phone')
        password = request.form.get('password')
        user = User.query.filter_by(phone=phone).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            if user.role == 'admin':
                return redirect('/admin')
            elif user.role == 'seller':
                return redirect('/seller')
            else:
                return redirect('/')
        else:
            flash('Invalid phone or password!')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.pop('referral', None)
    return redirect('/')

# ---------- PASSWORD RESET (FORGOT PASSWORD) ----------
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        if user:
            token = generate_reset_token(email)
            user.reset_token = token
            user.reset_token_expiry = datetime.utcnow() + timedelta(hours=1)
            db.session.commit()
            
            reset_url = url_for('reset_password', token=token, _external=True)
            subject = "Password Reset Request - ChoiceHub"
            body = f"""Hello {user.name},
You requested to reset your password for ChoiceHub.
Click the link below to reset your password (valid for 1 hour):
{reset_url}

If you did not request this, please ignore this email.
Regards,
ChoiceHub Team
"""
            msg = Message(subject, recipients=[email], body=body)
            try:
                print("📧 Sending email...")
                mail.send(msg)
                print("✅ Email sent successfully")
                flash('📧 Password reset link sent to your email. Please check your inbox (and spam folder).')
            except Exception as e:
                print("❌ MAIL ERROR:", str(e))
                flash(f'❌ Email sending failed: {str(e)}. Please try again later.')
        else:
            flash('❌ No account found with that email.')
        return redirect('/forgot-password')
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = User.query.filter_by(reset_token=token).first()
    if not user or user.reset_token_expiry < datetime.utcnow():
        flash('❌ Invalid or expired token. Please request a new reset link.')
        return redirect('/forgot-password')
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm = request.form.get('confirm_password')
        if password != confirm:
            flash('❌ Passwords do not match.')
            return redirect(request.url)
        if len(password) < 6:
            flash('❌ Password must be at least 6 characters.')
            return redirect(request.url)
        user.password_hash = generate_password_hash(password)
        user.reset_token = None
        user.reset_token_expiry = None
        db.session.commit()
        flash('✅ Password reset successful! Please login with your new password.')
        return redirect('/login')
    return render_template('reset_password.html', token=token)

# ---------- ROUTES (YOUR EXISTING CODE) ----------
# (All your existing routes go here – index, product, search, cart, checkout, profile, admin, etc.)
# I am pasting them below for completeness, but you can keep your own.

@app.route('/')
def index():
    ref = request.args.get('ref')
    if ref:
        session['referral'] = ref
    products = Product.query.filter_by(is_active=True).all()
    categories = Category.query.filter_by(is_active=True).all()
    banners = Banner.query.filter_by(is_active=True).order_by(Banner.position).all()
    featured = Product.query.filter_by(is_active=True, is_featured=True).limit(8).all()
    bestsellers = Product.query.filter_by(is_active=True, is_bestseller=True).limit(8).all()
    trending = Product.query.filter_by(is_active=True, is_trending=True).limit(8).all()
    new_arrivals = Product.query.filter_by(is_active=True, is_new=True).order_by(Product.created_at.desc()).limit(8).all()
    return render_template('index.html',
                           products=products,
                           categories=categories,
                           banners=banners,
                           featured=featured,
                           bestsellers=bestsellers,
                           trending=trending,
                           new_arrivals=new_arrivals)

@app.route('/product/<string:slug>')
def product_detail(slug):
    product = Product.query.filter_by(slug=slug, is_active=True).first_or_404()
    related = Product.query.filter(Product.category_id == product.category_id, Product.id != product.id, Product.is_active==True).limit(6).all()
    reviews = Review.query.filter_by(product_id=product.id, is_approved=True).order_by(Review.created_at.desc()).all()
    return render_template('product.html', product=product, related=related, reviews=reviews)

@app.route('/search')
def search():
    q = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    min_price = request.args.get('min_price', 0, type=float)
    max_price = request.args.get('max_price', 999999, type=float)
    sort = request.args.get('sort', 'newest')
    query = Product.query.filter_by(is_active=True)
    if q:
        query = query.filter(Product.name.contains(q) | Product.description.contains(q) | Product.brand.contains(q))
    if category:
        cat = Category.query.filter_by(slug=category).first()
        if cat:
            query = query.filter(Product.category_id == cat.id)
    query = query.filter(Product.selling_price >= min_price, Product.selling_price <= max_price)
    if sort == 'price_low':
        query = query.order_by(Product.selling_price.asc())
    elif sort == 'price_high':
        query = query.order_by(Product.selling_price.desc())
    else:
        query = query.order_by(Product.created_at.desc())
    products = query.all()
    categories = Category.query.filter_by(is_active=True).all()
    return render_template('search.html', products=products, q=q, categories=categories, selected_category=category)

@app.route('/check-pincode', methods=['POST'])
def check_pincode():
    data = request.json
    pincode = data.get('pincode')
    if not pincode or len(pincode) != 6 or not pincode.isdigit():
        return jsonify({'success': False, 'message': 'Invalid pincode'})
    delivery_date, days = estimate_delivery_date(pincode)
    if delivery_date:
        date_str = delivery_date.strftime('%A, %d %B')
        return jsonify({
            'success': True,
            'message': f'Delivery by {date_str}',
            'days': days,
            'cod_available': True,
            'free_delivery': True
        })
    else:
        return jsonify({'success': False, 'message': 'Delivery not available'})

# ---------- WISHLIST ----------
@app.route('/wishlist')
@login_required
def wishlist():
    if current_user.wishlist:
        ids = [int(x) for x in current_user.wishlist.split(',') if x.isdigit()]
        products = Product.query.filter(Product.id.in_(ids), Product.is_active==True).all()
    else:
        products = []
    return render_template('wishlist.html', products=products)

@app.route('/toggle-wishlist/<int:product_id>')
@login_required
def toggle_wishlist(product_id):
    product = Product.query.get_or_404(product_id)
    wish = current_user.wishlist or ''
    ids = [x for x in wish.split(',') if x.isdigit()]
    if str(product_id) in ids:
        ids.remove(str(product_id))
        flash('Removed from wishlist')
    else:
        ids.append(str(product_id))
        flash('Added to wishlist')
    current_user.wishlist = ','.join(ids)
    db.session.commit()
    return redirect(request.referrer or '/')

# ---------- CART ----------
@app.route('/add-to-cart/<int:product_id>', methods=['GET', 'POST'])
@login_required
def add_to_cart(product_id):
    if current_user.role != 'customer':
        flash('Only customers can buy.')
        return redirect('/')
    product = Product.query.get_or_404(product_id)
    qty = int(request.form.get('quantity', 1)) if request.method == 'POST' else 1
    if product.stock < qty:
        flash('Not enough stock!')
        return redirect('/product/' + product.slug)
    existing = Cart.query.filter_by(user_id=current_user.id, product_id=product_id).first()
    if existing:
        if existing.quantity + qty <= product.stock:
            existing.quantity += qty
        else:
            flash('Cannot add more than stock.')
            return redirect('/cart')
    else:
        cart = Cart(user_id=current_user.id, product_id=product_id, quantity=qty)
        db.session.add(cart)
    db.session.commit()
    flash('Added to cart!')
    return redirect('/cart')

@app.route('/cart')
@login_required
def view_cart():
    items = Cart.query.filter_by(user_id=current_user.id).all()
    total = cart_total(current_user.id)
    return render_template('cart.html', items=items, total=total)

@app.route('/update-cart/<int:cart_id>/<int:qty>')
@login_required
def update_cart(cart_id, qty):
    item = Cart.query.get_or_404(cart_id)
    if item.user_id != current_user.id:
        return "Unauthorized", 403
    if qty <= 0:
        db.session.delete(item)
    else:
        if qty <= item.product.stock:
            item.quantity = qty
        else:
            flash('Stock limit.')
    db.session.commit()
    return redirect('/cart')

@app.route('/remove-cart/<int:cart_id>')
@login_required
def remove_cart(cart_id):
    item = Cart.query.get_or_404(cart_id)
    if item.user_id == current_user.id:
        db.session.delete(item)
        db.session.commit()
    return redirect('/cart')

# ---------- COUPON ----------
@app.route('/apply-coupon', methods=['POST'])
@login_required
def apply_coupon():
    data = request.json
    code = data.get('code', '').strip().upper()
    total = float(data.get('total', 0))
    coupon = Coupon.query.filter_by(code=code, is_active=True).first()
    if not coupon:
        return jsonify({'success': False, 'message': 'Invalid coupon'})
    if datetime.utcnow() > coupon.expiry_date:
        return jsonify({'success': False, 'message': f'Expired on {coupon.expiry_date.strftime("%d %b %Y")}'})
    if total < coupon.min_order_amount:
        return jsonify({'success': False, 'message': f'Minimum order ₹{coupon.min_order_amount}'})
    discount = 0
    if coupon.discount_type == 'percentage':
        discount = (total * coupon.discount_value) / 100
        if coupon.max_discount_amount and discount > coupon.max_discount_amount:
            discount = coupon.max_discount_amount
    else:
        discount = coupon.discount_value
    net = total - discount
    return jsonify({
        'success': True,
        'discount': round(discount, 2),
        'net': round(net, 2),
        'coupon_id': coupon.id
    })

# ---------- CHECKOUT ----------
@app.route('/checkout', methods=['GET', 'POST'])
@login_required
def checkout():
    if current_user.role != 'customer':
        return redirect('/')
    items = Cart.query.filter_by(user_id=current_user.id).all()
    if not items:
        return redirect('/')
    total = cart_total(current_user.id)
    addresses = Address.query.filter_by(user_id=current_user.id).all()

    if request.method == 'POST':
        address_id = request.form.get('address_id')
        if address_id:
            addr = Address.query.get_or_404(address_id)
            full_name = addr.full_name
            phone = addr.phone
            address_line1 = addr.address_line1
            address_line2 = addr.address_line2 or ''
            locality = addr.locality or ''
            landmark = addr.landmark or ''
            city = addr.city
            state = addr.state
            pincode = addr.pincode
        else:
            full_name = request.form.get('full_name')
            phone = request.form.get('phone')
            address_line1 = request.form.get('address_line1')
            address_line2 = request.form.get('address_line2', '')
            locality = request.form.get('locality', '')
            landmark = request.form.get('landmark', '')
            city = request.form.get('city')
            state = request.form.get('state')
            pincode = request.form.get('pincode')

        # Validate
        if not all([full_name, phone, address_line1, city, state, pincode]):
            flash('Please fill all required fields.')
            return redirect('/checkout')
        if not pincode.isdigit() or len(pincode) != 6:
            flash('Enter a valid 6-digit pincode.')
            return redirect('/checkout')

        # Build address text
        address_text = f"{full_name}, {phone}, {address_line1}"
        if address_line2:
            address_text += f", {address_line2}"
        if locality:
            address_text += f", {locality}"
        if landmark:
            address_text += f", near {landmark}"
        address_text += f", {city}, {state} - {pincode}"

        # Save address if new
        if not address_id:
            new_addr = Address(
                user_id=current_user.id,
                full_name=full_name,
                phone=phone,
                address_line1=address_line1,
                address_line2=address_line2,
                city=city,
                state=state,
                pincode=pincode,
                locality=locality,
                landmark=landmark,
                is_default=not Address.query.filter_by(user_id=current_user.id).first()
            )
            db.session.add(new_addr)
            db.session.flush()

        coupon_id = request.form.get('coupon_id')
        # Safely convert discount and net
        discount_str = request.form.get('discount', '0').strip()
        discount = float(discount_str) if discount_str else 0.0
        net_str = request.form.get('net', str(total)).strip()
        net = float(net_str) if net_str else total

        order_num = generate_order_number()

        try:
            delivery_date, days = estimate_delivery_date(pincode)
            order = Order(
                order_number=order_num,
                user_id=current_user.id,
                coupon_id=int(coupon_id) if coupon_id else None,
                referral_used=session.get('referral'),
                total_amount=total,
                discount_amount=discount,
                shipping_charge=0.0,
                net_amount=net,
                shipping_address=address_text,
                pincode=pincode,
                delivery_date=delivery_date,
                payment_method='COD'
            )
            db.session.add(order)
            db.session.flush()
            for item in items:
                seller = Seller.query.filter_by(user_id=item.product.seller_id).first()
                seller_id = seller.id if seller else None
                if seller and session.get('referral') == seller.user.referral_code:
                    seller.total_earned += (item.product.selling_price * item.quantity * (seller.commission_rate / 100))
                oi = OrderItem(
                    order_id=order.id,
                    product_id=item.product_id,
                    seller_id=seller_id,
                    quantity=item.quantity,
                    price=item.product.selling_price
                )
                db.session.add(oi)
                product = Product.query.get(item.product_id)
                product.stock -= item.quantity
                db.session.delete(item)
            db.session.commit()
            flash(f'✅ Order placed! Order ID: {order_num}')
            return redirect('/profile')
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error placing order: {str(e)}')
            return redirect('/checkout')

    return render_template('checkout.html', items=items, total=total, addresses=addresses)

# ---------- PROFILE & ADDRESS MANAGEMENT ----------
@app.route('/profile')
@login_required
def profile():
    orders = Order.query.filter_by(user_id=current_user.id).order_by(Order.created_at.desc()).all()
    return render_template('profile.html', orders=orders)

@app.route('/profile/addresses')
@login_required
def manage_addresses():
    addresses = Address.query.filter_by(user_id=current_user.id).all()
    return render_template('addresses.html', addresses=addresses)

@app.route('/profile/addresses/add', methods=['POST'])
@login_required
def add_address():
    full_name = request.form.get('full_name')
    phone = request.form.get('phone')
    address_line1 = request.form.get('address_line1')
    address_line2 = request.form.get('address_line2')
    locality = request.form.get('locality')
    landmark = request.form.get('landmark')
    city = request.form.get('city')
    state = request.form.get('state')
    pincode = request.form.get('pincode')
    if not all([full_name, phone, address_line1, city, state, pincode]):
        flash('Please fill all required fields.')
        return redirect('/profile/addresses')
    if not pincode.isdigit() or len(pincode) != 6:
        flash('Enter a valid 6-digit pincode.')
        return redirect('/profile/addresses')
    addr = Address(
        user_id=current_user.id,
        full_name=full_name,
        phone=phone,
        address_line1=address_line1,
        address_line2=address_line2,
        locality=locality,
        landmark=landmark,
        city=city,
        state=state,
        pincode=pincode,
        is_default=not Address.query.filter_by(user_id=current_user.id).first()
    )
    db.session.add(addr)
    db.session.commit()
    flash('Address added!')
    return redirect('/profile/addresses')

@app.route('/profile/addresses/delete/<int:id>')
@login_required
def delete_address(id):
    addr = Address.query.get_or_404(id)
    if addr.user_id != current_user.id:
        return "Unauthorized", 403
    db.session.delete(addr)
    db.session.commit()
    flash('Address deleted.')
    return redirect('/profile/addresses')

@app.route('/cancel-order/<int:order_id>')
@login_required
def cancel_order(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id == current_user.id and order.status in ['Placed', 'Confirmed']:
        order.status = 'Cancelled'
        db.session.commit()
        flash('Order cancelled.')
    return redirect('/profile')

@app.route('/add-review/<int:product_id>', methods=['POST'])
@login_required
def add_review(product_id):
    rating = int(request.form.get('rating'))
    comment = request.form.get('comment')
    title = request.form.get('title')
    order_item = OrderItem.query.join(Order).filter(Order.user_id==current_user.id, OrderItem.product_id==product_id).first()
    verified = True if order_item else False
    review = Review(
        user_id=current_user.id,
        product_id=product_id,
        rating=rating,
        title=title,
        comment=comment,
        is_verified=verified,
        is_approved=True
    )
    db.session.add(review)
    db.session.commit()
    flash('Review submitted!')
    return redirect('/product/' + Product.query.get(product_id).slug)

# ---------- ADMIN PANEL (UPDATED WITH TAB PERSISTENCY) ----------
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return "Access Denied", 403
    products = Product.query.all()
    orders = Order.query.all()
    users = User.query.all()
    coupons = Coupon.query.all()
    sellers = Seller.query.all()
    categories = Category.query.all()
    reviews = Review.query.all()
    banners = Banner.query.all()
    stats = {
        'total_orders': len(orders),
        'total_products': len(products),
        'total_users': len(users),
        'total_revenue': sum(o.net_amount for o in orders if o.status == 'Delivered'),
        'pending_orders': len([o for o in orders if o.status == 'Placed']),
        'low_stock': len([p for p in products if p.stock < 5]),
    }
    return render_template('admin.html',
                           products=products,
                           orders=orders,
                           users=users,
                           coupons=coupons,
                           sellers=sellers,
                           categories=categories,
                           reviews=reviews,
                           banners=banners,
                           stats=stats)

@app.route('/admin/add-product', methods=['POST'])
@login_required
def admin_add_product():
    if current_user.role != 'admin':
        return "Access Denied", 403
    name = request.form.get('name')
    slug = request.form.get('slug') or name.lower().replace(' ', '-')
    if Product.query.filter_by(slug=slug).first():
        slug = slug + '-' + uuid.uuid4().hex[:4]
    category_id = int(request.form.get('category_id'))
    brand = request.form.get('brand')
    sku = request.form.get('sku') or f"SKU-{uuid.uuid4().hex[:8].upper()}"
    cost_price = float(request.form.get('cost_price', 0))
    selling_price = float(request.form.get('selling_price'))
    mrp = float(request.form.get('mrp'))
    stock = int(request.form.get('stock', 0))
    short_desc = request.form.get('short_description')
    description = request.form.get('description')
    features = request.form.get('features')
    specifications = request.form.get('specifications')
    weight = request.form.get('weight', type=float)
    dimensions = request.form.get('dimensions')
    material = request.form.get('material')
    care_instructions = request.form.get('care_instructions')
    warranty = request.form.get('warranty')
    return_policy = request.form.get('return_policy')
    delivery_time = request.form.get('delivery_time')
    video_url = request.form.get('video_url')
    is_featured = request.form.get('is_featured') == 'on'
    is_bestseller = request.form.get('is_bestseller') == 'on'
    is_trending = request.form.get('is_trending') == 'on'
    is_new = request.form.get('is_new') == 'on'
    free_delivery = request.form.get('free_delivery') == 'on'
    cod_available = request.form.get('cod_available') == 'on'
    seller_id = request.form.get('seller_id')
    files = request.files.getlist('images')
    fnames = []
    for f in files:
        if f and allowed_file(f.filename):
            fn = upload_image(f)
            if fn:
                fnames.append(fn)
    images_str = ','.join(fnames) if fnames else ''
    product = Product(
        name=name, slug=slug, category_id=category_id,
        brand=brand, sku=sku, cost_price=cost_price,
        selling_price=selling_price, mrp=mrp,
        stock=stock, short_description=short_desc,
        description=description, features=features,
        specifications=specifications, weight=weight,
        dimensions=dimensions, material=material,
        care_instructions=care_instructions,
        warranty=warranty, return_policy=return_policy,
        delivery_time=delivery_time, video_url=video_url,
        is_featured=is_featured, is_bestseller=is_bestseller,
        is_trending=is_trending, is_new=is_new,
        free_delivery=free_delivery, cod_available=cod_available,
        seller_id=int(seller_id) if seller_id else None,
        images=images_str
    )
    if mrp > 0:
        product.discount_percent = ((mrp - selling_price) / mrp) * 100
    db.session.add(product)
    db.session.commit()
    flash('✅ Product added!')
    return redirect('/admin?tab=products')

@app.route('/admin/edit-product/<int:id>', methods=['GET', 'POST'])
@login_required
def admin_edit_product(id):
    if current_user.role != 'admin':
        return "Access Denied", 403
    product = Product.query.get_or_404(id)
    if request.method == 'POST':
        product.name = request.form.get('name')
        product.slug = request.form.get('slug') or product.name.lower().replace(' ', '-')
        product.category_id = int(request.form.get('category_id'))
        product.brand = request.form.get('brand')
        product.sku = request.form.get('sku')
        product.cost_price = float(request.form.get('cost_price', 0))
        product.selling_price = float(request.form.get('selling_price'))
        product.mrp = float(request.form.get('mrp'))
        product.stock = int(request.form.get('stock', 0))
        product.short_description = request.form.get('short_description')
        product.description = request.form.get('description')
        product.features = request.form.get('features')
        product.specifications = request.form.get('specifications')
        product.weight = request.form.get('weight', type=float)
        product.dimensions = request.form.get('dimensions')
        product.material = request.form.get('material')
        product.care_instructions = request.form.get('care_instructions')
        product.warranty = request.form.get('warranty')
        product.return_policy = request.form.get('return_policy')
        product.video_url = request.form.get('video_url')
        product.is_featured = request.form.get('is_featured') == 'on'
        product.is_bestseller = request.form.get('is_bestseller') == 'on'
        product.is_trending = request.form.get('is_trending') == 'on'
        product.is_new = request.form.get('is_new') == 'on'
        product.free_delivery = request.form.get('free_delivery') == 'on'
        product.cod_available = request.form.get('cod_available') == 'on'
        product.seller_id = request.form.get('seller_id')
        if product.mrp > 0:
            product.discount_percent = ((product.mrp - product.selling_price) / product.mrp) * 100
        db.session.commit()
        flash('✅ Product updated!')
        return redirect('/admin?tab=products')
    categories = Category.query.all()
    sellers = Seller.query.all()
    return render_template('edit_product.html', product=product, categories=categories, sellers=sellers)

@app.route('/admin/delete-product/<int:id>')
@login_required
def admin_delete_product(id):
    if current_user.role != 'admin':
        return "Access Denied", 403
    product = Product.query.get_or_404(id)
    for img in product.image_list():
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], img))
        except:
            pass
    db.session.delete(product)
    db.session.commit()
    return redirect('/admin?tab=products')

@app.route('/admin/update-order/<int:id>/<status>')
@login_required
def admin_update_order(id, status):
    if current_user.role != 'admin':
        return "Access Denied", 403
    order = Order.query.get_or_404(id)
    order.status = status
    db.session.commit()
    return redirect('/admin?tab=orders')

@app.route('/admin/add-coupon', methods=['POST'])
@login_required
def admin_add_coupon():
    if current_user.role != 'admin':
        return "Access Denied", 403
    code = request.form.get('code').upper()
    dtype = request.form.get('discount_type')
    dvalue = float(request.form.get('discount_value'))
    min_order = float(request.form.get('min_order', 0))
    max_discount = request.form.get('max_discount')
    max_discount = float(max_discount) if max_discount else None
    expiry_str = request.form.get('expiry')
    expiry = datetime.strptime(expiry_str, '%Y-%m-%d') if expiry_str else datetime.utcnow() + timedelta(days=30)
    coupon = Coupon(
        code=code,
        discount_type=dtype,
        discount_value=dvalue,
        min_order_amount=min_order,
        max_discount_amount=max_discount,
        expiry_date=expiry
    )
    db.session.add(coupon)
    db.session.commit()
    flash(f'✅ Coupon {code} added!')
    return redirect('/admin?tab=coupons')

@app.route('/admin/delete-coupon/<int:id>')
@login_required
def admin_delete_coupon(id):
    if current_user.role != 'admin':
        return "Access Denied", 403
    coupon = Coupon.query.get_or_404(id)
    db.session.delete(coupon)
    db.session.commit()
    return redirect('/admin?tab=coupons')

@app.route('/admin/create-seller', methods=['POST'])
@login_required
def admin_create_seller():
    if current_user.role != 'admin':
        return "Access Denied", 403
    name = request.form.get('name')
    phone = request.form.get('phone')
    password = request.form.get('password')
    store_name = request.form.get('store_name')
    commission = float(request.form.get('commission_rate', 10))
    if User.query.filter_by(phone=phone).first():
        flash('Phone already registered!')
        return redirect('/admin?tab=sellers')
    hashed = generate_password_hash(password)
    ref_code = f"{name[:4].upper()}{random.randint(100,999)}"
    while User.query.filter_by(referral_code=ref_code).first():
        ref_code = f"{name[:4].upper()}{random.randint(100,999)}"
    user = User(
        name=name,
        phone=phone,
        password_hash=hashed,
        role='seller',
        referral_code=ref_code
    )
    db.session.add(user)
    db.session.flush()
    seller = Seller(user_id=user.id, store_name=store_name, commission_rate=commission)
    db.session.add(seller)
    db.session.commit()
    flash(f'✅ Seller {name} created! Phone: {phone}, Password: {password}')
    return redirect('/admin?tab=sellers')

@app.route('/admin/import-excel', methods=['POST'])
@login_required
def admin_import_excel():
    if current_user.role != 'admin':
        return "Access Denied", 403
    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('Please upload .xlsx file')
        return redirect('/admin?tab=products')
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        col_map = {}
        for idx, h in enumerate(headers):
            h_lower = str(h).strip().lower()
            if h_lower in ['name', 'brand', 'sku', 'cost_price', 'selling_price', 'mrp', 'stock',
                           'category', 'short_description', 'description', 'features', 'specifications',
                           'weight', 'dimensions', 'material', 'care_instructions', 'warranty',
                           'return_policy', 'delivery_time', 'video_url', 'image_urls',
                           'is_featured', 'is_bestseller', 'is_trending', 'is_new',
                           'free_delivery', 'cod_available']:
                col_map[h_lower] = idx
        required = ['name', 'selling_price', 'mrp']
        for req in required:
            if req not in col_map:
                flash(f'Missing column: {req}')
                return redirect('/admin?tab=products')
        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_map['name']]:
                continue
            name = str(row[col_map['name']])
            selling_price = float(row[col_map['selling_price']])
            mrp = float(row[col_map['mrp']])
            cost_price = float(row[col_map.get('cost_price', -1)]) if 'cost_price' in col_map and row[col_map['cost_price']] is not None else 0
            stock = int(row[col_map.get('stock', -1)]) if 'stock' in col_map and row[col_map['stock']] is not None else 0
            category_name = str(row[col_map.get('category', -1)]) if 'category' in col_map and row[col_map['category']] else ''
            brand = str(row[col_map.get('brand', -1)]) if 'brand' in col_map and row[col_map['brand']] else ''
            sku = str(row[col_map.get('sku', -1)]) if 'sku' in col_map and row[col_map['sku']] else f"SKU-{uuid.uuid4().hex[:8].upper()}"
            short_desc = str(row[col_map.get('short_description', -1)]) if 'short_description' in col_map and row[col_map['short_description']] else ''
            description = str(row[col_map.get('description', -1)]) if 'description' in col_map and row[col_map['description']] else ''
            features = str(row[col_map.get('features', -1)]) if 'features' in col_map and row[col_map['features']] else ''
            specifications = str(row[col_map.get('specifications', -1)]) if 'specifications' in col_map and row[col_map['specifications']] else ''
            weight = float(row[col_map.get('weight', -1)]) if 'weight' in col_map and row[col_map['weight']] is not None else None
            dimensions = str(row[col_map.get('dimensions', -1)]) if 'dimensions' in col_map and row[col_map['dimensions']] else ''
            material = str(row[col_map.get('material', -1)]) if 'material' in col_map and row[col_map['material']] else ''
            care_instructions = str(row[col_map.get('care_instructions', -1)]) if 'care_instructions' in col_map and row[col_map['care_instructions']] else ''
            warranty = str(row[col_map.get('warranty', -1)]) if 'warranty' in col_map and row[col_map['warranty']] else ''
            return_policy = str(row[col_map.get('return_policy', -1)]) if 'return_policy' in col_map and row[col_map['return_policy']] else ''
            delivery_time = str(row[col_map.get('delivery_time', -1)]) if 'delivery_time' in col_map and row[col_map['delivery_time']] else ''
            video_url = str(row[col_map.get('video_url', -1)]) if 'video_url' in col_map and row[col_map['video_url']] else ''
            image_urls = str(row[col_map.get('image_urls', -1)]) if 'image_urls' in col_map and row[col_map['image_urls']] else ''
            is_featured = str(row[col_map.get('is_featured', -1)]) if 'is_featured' in col_map and row[col_map['is_featured']] else ''
            is_bestseller = str(row[col_map.get('is_bestseller', -1)]) if 'is_bestseller' in col_map and row[col_map['is_bestseller']] else ''
            is_trending = str(row[col_map.get('is_trending', -1)]) if 'is_trending' in col_map and row[col_map['is_trending']] else ''
            is_new = str(row[col_map.get('is_new', -1)]) if 'is_new' in col_map and row[col_map['is_new']] else ''
            free_delivery = str(row[col_map.get('free_delivery', -1)]) if 'free_delivery' in col_map and row[col_map['free_delivery']] else ''
            cod_available = str(row[col_map.get('cod_available', -1)]) if 'cod_available' in col_map and row[col_map['cod_available']] else ''
            # Category
            cat = Category.query.filter_by(name=category_name).first()
            if not cat and category_name:
                cat = Category(name=category_name, slug=category_name.lower().replace(' ', '-'))
                db.session.add(cat)
                db.session.flush()
            elif not cat:
                cat = Category.query.first()
            # Images
            images_list = []
            if image_urls:
                for url in image_urls.split(','):
                    url = url.strip()
                    if url:
                        fn = download_and_save_image(url)
                        if fn:
                            images_list.append(fn)
            images_str = ','.join(images_list)
            slug = name.lower().replace(' ', '-') + '-' + uuid.uuid4().hex[:4]
            product = Product(
                name=name, slug=slug, category_id=cat.id,
                brand=brand, sku=sku, cost_price=cost_price,
                selling_price=selling_price, mrp=mrp,
                stock=stock, short_description=short_desc,
                description=description, features=features,
                specifications=specifications, weight=weight,
                dimensions=dimensions, material=material,
                care_instructions=care_instructions,
                warranty=warranty, return_policy=return_policy,
                delivery_time=delivery_time, video_url=video_url,
                is_featured=(is_featured.lower() in ['true','yes','1']) if is_featured else False,
                is_bestseller=(is_bestseller.lower() in ['true','yes','1']) if is_bestseller else False,
                is_trending=(is_trending.lower() in ['true','yes','1']) if is_trending else False,
                is_new=(is_new.lower() in ['true','yes','1']) if is_new else False,
                free_delivery=(free_delivery.lower() in ['true','yes','1']) if free_delivery else False,
                cod_available=(cod_available.lower() in ['true','yes','1']) if cod_available else False,
                images=images_str
            )
            if mrp > 0:
                product.discount_percent = ((mrp - selling_price) / mrp) * 100
            db.session.add(product)
            added += 1
        db.session.commit()
        flash(f'✅ Imported {added} products!')
    except Exception as e:
        flash(f'Error: {str(e)}')
    return redirect('/admin?tab=products')

@app.route('/admin/download-template')
@login_required
def download_template():
    if current_user.role != 'admin':
        return "Access Denied", 403
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = [
        'name', 'brand', 'sku', 'cost_price', 'selling_price', 'mrp', 'stock',
        'category', 'short_description', 'description', 'features', 'specifications',
        'weight', 'dimensions', 'material', 'care_instructions', 'warranty',
        'return_policy', 'delivery_time', 'video_url', 'image_urls',
        'is_featured', 'is_bestseller', 'is_trending', 'is_new',
        'free_delivery', 'cod_available'
    ]
    ws.append(headers)
    sample = [
        'Sample Product', 'SampleBrand', 'SKU123', 300, 499, 699, 10,
        'Gifts', 'Short description', 'Full description here', 'Feature1, Feature2', 'Weight: 1kg, Color: Gold',
        1.2, '10x10x5 cm', 'Wood', 'Wipe with dry cloth', '1 year', '7 days return', '2-3 days', 'https://youtube.com/xyz',
        'https://example.com/image1.jpg, https://example.com/image2.jpg  (comma separated URLs)',
        'yes', 'no', 'yes', 'no', 'yes', 'yes'
    ]
    ws.append(sample)
    from openpyxl.comments import Comment
    cell = ws['V2']
    cell.comment = Comment('Enter comma-separated image URLs. The system will download and save them automatically.', 'Admin')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='product_import_template.xlsx', as_attachment=True)

@app.route('/admin/add-category', methods=['POST'])
@login_required
def admin_add_category():
    if current_user.role != 'admin':
        return "Access Denied", 403
    name = request.form.get('name')
    slug = name.lower().replace(' ', '-')
    if Category.query.filter_by(slug=slug).first():
        flash('Category already exists!')
        return redirect('/admin?tab=categories')
    cat = Category(name=name, slug=slug)
    db.session.add(cat)
    db.session.commit()
    flash('✅ Category added!')
    return redirect('/admin?tab=categories')

@app.route('/admin/delete-category/<int:id>')
@login_required
def admin_delete_category(id):
    if current_user.role != 'admin':
        return "Access Denied", 403
    cat = Category.query.get_or_404(id)
    db.session.delete(cat)
    db.session.commit()
    return redirect('/admin?tab=categories')

@app.route('/admin/add-banner', methods=['POST'])
@login_required
def admin_add_banner():
    if current_user.role != 'admin':
        return "Access Denied", 403
    title = request.form.get('title')
    link = request.form.get('link')
    file = request.files.get('image')
    if file and allowed_file(file.filename):
        filename = upload_image(file)
        banner = Banner(title=title, image=filename, link=link)
        db.session.add(banner)
        db.session.commit()
        flash('✅ Banner added!')
    else:
        flash('Please upload a valid image.')
    return redirect('/admin?tab=banners')

@app.route('/admin/delete-banner/<int:id>')
@login_required
def admin_delete_banner(id):
    if current_user.role != 'admin':
        return "Access Denied", 403
    banner = Banner.query.get_or_404(id)
    try:
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], banner.image))
    except:
        pass
    db.session.delete(banner)
    db.session.commit()
    return redirect('/admin?tab=banners')

# ---------- BUY NOW ----------
@app.route('/buy-now/<int:product_id>', methods=['POST'])
@login_required
def buy_now(product_id):
    if current_user.role != 'customer':
        flash('Only customers can buy.')
        return redirect('/')
    product = Product.query.get_or_404(product_id)
    if product.stock <= 0:
        flash('Out of stock!')
        return redirect('/product/' + product.slug)
    existing = Cart.query.filter_by(user_id=current_user.id, product_id=product_id).first()
    if not existing:
        cart = Cart(user_id=current_user.id, product_id=product_id, quantity=1)
        db.session.add(cart)
        db.session.commit()
    flash('Product added to cart. Proceed to checkout.')
    return redirect('/checkout')

# ---------- SELLER PANEL ----------
@app.route('/seller')
@login_required
def seller_dashboard():
    if current_user.role != 'seller':
        return "Access Denied", 403
    seller = Seller.query.filter_by(user_id=current_user.id).first()
    if not seller:
        flash('No seller profile.')
        return redirect('/')
    products = Product.query.filter_by(seller_id=seller.id).all()
    order_items = OrderItem.query.filter_by(seller_id=seller.id).all()
    order_ids = [oi.order_id for oi in order_items]
    orders = Order.query.filter(Order.id.in_(order_ids)).order_by(Order.created_at.desc()).all() if order_ids else []
    ref_link = f"{request.host_url}?ref={current_user.referral_code}"
    return render_template('seller.html', seller=seller, products=products, orders=orders, ref_link=ref_link)

# ---------- ERROR HANDLING ----------
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template('500.html'), 500

# ---------- RUN ----------
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)   # production mode
